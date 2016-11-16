import openpyxl
import random
from openpyxl.chart import Reference, Series, LineChart

wb = openpyxl.load_workbook('sample.xlsx')
print(wb.get_sheet_names())

ws = wb.get_sheet_by_name('Sheet1')
print(ws.max_column)
print(ws.max_row)

print(ws['A4'].value)
a3 = ws.cell(row=3, column=1)
print(a3.value)
b5 = ws.cell('B5')
print(b5.value)

wb2 = openpyxl.load_workbook('sample.xlsx', data_only=True)
ws2 = wb2.get_active_sheet()

for row in ws2.rows:
    for cell in row:
        print(cell.value)

for column in ws2.columns:
    for cell in column:
        print(cell.value)

wb3 = openpyxl.Workbook()
ws3 = wb3.create_sheet(index=0, title='New Sheet')
ws3['A1'] = 100

wb3.save(filename='new_book.xlsx')

wb4 = openpyxl.Workbook()
ws4 = wb4.active

for i in range(10):
    ws4.append([random.randint(1, 10)])

values = Reference(ws4, min_row=1, min_col=1, max_row=10, max_col=1)
series = Series(values, title='Sample Chart')
chart = LineChart()
chart.append(series)
ws4.add_chart(chart)
wb4.save('sample_chart.xlsx')
